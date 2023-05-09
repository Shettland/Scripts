import os
import sys
import json
import openpyxl

"""
A script built to remove corruped samples from both excel and json metadata

Usage: script.py excel_file json_file corrupted_samples_file

The file containing the corrupted sample names should have one sample name
per line, like so:

AND01301
AND01395
MCN0714
"""

"""
shell command to remove the sample files from the folder
while IFS= read -r file; do cp "folder/with/samples/${file}.R"* output/location/ ; done < corrupted_samples_file.txt
while IFS= read -r file; do rm -- "folder/with/samples/${file}.R"* ; done < corrupted_samples_file.txt
"""

args = sys.argv

if len(args) == 4:
    excel_path = args[1]
    json_path = args[2]
    corrupted_samples_file = args[3]
    if not os.path.exists(excel_path):
        print("Could not find excel file")
        sys.exit(1)
    if not os.path.exists(json_path):
        print("Could not find json file")
        sys.exit(1)
    if not os.path.exists(corrupted_samples_file):
        print("Could not find corruped samples file")
        sys.exit(1)
else:
    print("The script needs 3 arguments:")
    print("Usage: script.py excel_file json_file corrupted_samples_file")
    sys.exit(1)

def clean_excel_rows(excel_path, json_path, corrupted_samples_file):
    excel_filepath = excel_path
    excel_dirname = os.path.dirname(excel_filepath)
    excel_filename = os.path.basename(excel_filepath)
    wb_file = openpyxl.load_workbook(excel_filepath, data_only=True)
    ws_metadata_lab = wb_file["METADATA_LAB"]

    with open(corrupted_samples_file) as f:
        lines = f.readlines()
    lines = [s.strip('\n') for s in lines]
    row_indexes = []

    print("Cleaning json data...")
    clean_json_data(json_path, lines)

    for row in ws_metadata_lab.iter_rows(min_row=1, max_row=ws_metadata_lab.max_row):
        flag = False
        for cell in row:
            for target_string in lines:
                if target_string in str(cell.value):
                    row_indexes.append(cell.row)
                    flag = True
                    break
            if flag:
                break

    test_labels=[]
    for index in row_indexes:
        label = [v.value for v in ws_metadata_lab[index]][3]
        test_labels.append(label)
    if test_labels == lines:
        print("Correctly validated rows to delete, proceeding...")
        for row_idx in sorted(row_indexes, reverse=True):
            ws_metadata_lab.delete_rows(row_idx)
    else:
        print("Could not find the correct rows to delete, aborting")
        sys.exit(1)

    clean_excelname = str("clean_"+excel_filename)
    output_excel_path = os.path.join(excel_dirname, clean_excelname)
    wb_file.save(output_excel_path)
    print("Correctly saved excel file as", output_excel_path)
    return

def clean_json_data(json_path, lines):
    sample_data_path = json_path

    with open(sample_data_path, "r", encoding="utf-8") as fh:
        json_data = json.load(fh)

    for corrupted_sample in lines:
        try:
            del json_data[corrupted_sample]
        except KeyError as e:
            print(f"Sample {e} could not be found in json file")
            continue

    json_dirname = os.path.dirname(sample_data_path)
    json_filename = os.path.basename(sample_data_path)
    clean_jsonname = str("clean_"+json_filename)
    output_json_path = os.path.join(json_dirname,clean_jsonname)
    
    with open(output_json_path, "w", encoding="utf-8") as fh:
        fh.write(json.dumps(json_data, indent=4, sort_keys=True, ensure_ascii=False))
    print("Correctly saved json file as", output_json_path)

    return

clean_excel_rows(excel_path, json_path, corrupted_samples_file)
